import openpyxl
import sys

# 讀取Excel檔案
file_path = r"C:\Users\USER\Desktop\更動1.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
    print(f"成功讀取檔案: {file_path}")
    print(f"\n工作表名稱: {wb.sheetnames}")

    # 讀取第一個工作表
    ws = wb.active
    print(f"\n當前工作表: {ws.title}")
    print(f"最大行數: {ws.max_row}")
    print(f"最大列數: {ws.max_column}")

    # 顯示前10行資料
    print("\n前10行資料預覽：")
    print("-" * 100)
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 10:
            break
        print(f"Row {i}: {row}")

except Exception as e:
    print(f"錯誤: {e}")
    sys.exit(1)
