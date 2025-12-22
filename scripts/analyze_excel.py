#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel 資料分析工具
讀取「更動1.xlsx」並分析資料結構
"""

import openpyxl
from collections import defaultdict
import sys

def analyze_excel(filepath):
    """分析 Excel 檔案的資料結構"""

    print(f"正在讀取檔案: {filepath}\n")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)

        print(f"工作表總數: {len(wb.sheetnames)}")
        print(f"工作表清單: {wb.sheetnames}\n")
        print("=" * 80)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            print(f"\n【工作表: {sheet_name}】")
            print(f"最大行數: {ws.max_row}")
            print(f"最大列數: {ws.max_column}")

            # 讀取標題列
            if ws.max_row > 0:
                headers = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    headers.append(cell_value if cell_value else f"<空白{col}>")

                print(f"\n標題列 ({len(headers)} 欄):")
                for idx, header in enumerate(headers, 1):
                    print(f"  {idx}. {header}")

            # 如果是 sheet3，分析公司分布
            if sheet_name == "sheet3" and ws.max_row > 1:
                print("\n【資料分析】")

                # 尋找「資料來源」和「公司名稱」欄位
                source_col = None
                company_col = None

                for col in range(1, ws.max_column + 1):
                    header = ws.cell(row=1, column=col).value
                    if header and "資料來源" in str(header):
                        source_col = col
                    if header and "公司名稱" in str(header):
                        company_col = col

                if source_col and company_col:
                    # 統計資料來源分布
                    source_count = defaultdict(int)
                    company_records = defaultdict(int)

                    for row in range(2, ws.max_row + 1):
                        source = ws.cell(row=row, column=source_col).value
                        company = ws.cell(row=row, column=company_col).value

                        if source:
                            source_count[source] += 1
                        if company:
                            company_records[company] += 1

                    print(f"\n資料來源分布:")
                    for source, count in sorted(source_count.items()):
                        print(f"  {source}: {count} 筆")

                    print(f"\n唯一公司數: {len(company_records)}")
                    print(f"總資料筆數: {sum(company_records.values())}")

                    # 前 10 大公司
                    top_companies = sorted(company_records.items(),
                                         key=lambda x: x[1],
                                         reverse=True)[:10]

                    print("\n前 10 大公司（依記錄數）:")
                    for rank, (company, count) in enumerate(top_companies, 1):
                        print(f"  {rank}. {company}: {count} 筆")

            print("\n" + "-" * 80)

        wb.close()
        print("\n✅ 分析完成")

    except FileNotFoundError:
        print(f"❌ 錯誤: 找不到檔案 {filepath}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ 錯誤: {e}")
        sys.exit(1)

if __name__ == "__main__":
    filepath = r"C:\Users\USER\Desktop\更動1.xlsx"
    analyze_excel(filepath)
